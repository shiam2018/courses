#!/usr/bin/env python3
"""Build the downloadable practice workbook for course 183094 (free Excel
course). One realistic marketing-campaigns dataset reused across sheets, plus
a lookup table for VLOOKUP/INDEX+MATCH practice and an exercises sheet.
"""
import random
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(42)

CHANNELS = ["Яндекс.Директ", "VK Реклама", "SEO", "Email", "Партнёрка"]
CHANNEL_CODE = {"Яндекс.Директ": "YD", "VK Реклама": "VK", "SEO": "SEO", "Email": "EM", "Партнёрка": "PT"}
STATUSES = ["Активна", "Приостановлена", "Завершена"]

BRAND_HEADER_FILL = PatternFill("solid", fgColor="E2532F")
BRAND_HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1C1A18")
NOTE_FONT = Font(italic=True, color="57524B")


def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = BRAND_HEADER_FILL
        c.font = BRAND_HEADER_FONT
        c.alignment = Alignment(horizontal="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_data_sheet(wb):
    ws = wb.active
    ws.title = "Данные"
    ws["A1"] = "Рекламные кампании — тренировочный датасет"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Используйте эту таблицу для формул, ВПР, сортировки, фильтров и сводных таблиц."
    ws["A2"].font = NOTE_FONT

    headers = ["Дата", "Канал", "Код канала", "Кампания", "Расход, ₽", "Показы", "Клики", "Заявки", "Статус"]
    header_row = 4
    for j, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=j, value=h)
    style_header_row(ws, header_row, len(headers))

    start = date(2026, 1, 1)
    row = header_row + 1
    for i in range(180):
        d = start + timedelta(days=random.randint(0, 210))
        channel = random.choice(CHANNELS)
        spend = round(random.uniform(500, 15000), 0)
        impressions = random.randint(500, 50000)
        ctr = random.uniform(0.005, 0.06)
        clicks = max(1, int(impressions * ctr))
        cr = random.uniform(0.01, 0.12)
        leads = int(clicks * cr)
        status = random.choices(STATUSES, weights=[0.6, 0.15, 0.25])[0]
        campaign_name = f"{CHANNEL_CODE[channel]}_{d.strftime('%Y%m')}_{i % 12 + 1:02d}"
        ws.cell(row=row, column=1, value=d).number_format = "DD.MM.YYYY"
        ws.cell(row=row, column=2, value=channel)
        ws.cell(row=row, column=3, value=CHANNEL_CODE[channel])
        ws.cell(row=row, column=4, value=campaign_name)
        ws.cell(row=row, column=5, value=spend)
        ws.cell(row=row, column=6, value=impressions)
        ws.cell(row=row, column=7, value=clicks)
        ws.cell(row=row, column=8, value=leads)
        ws.cell(row=row, column=9, value=status)
        row += 1

    autosize(ws, [12, 16, 10, 16, 12, 10, 10, 10, 14])
    ws.freeze_panes = "A5"
    return ws


def build_lookup_sheet(wb):
    ws = wb.create_sheet("Справочник каналов")
    ws["A1"] = "Справочник для ВПР / ИНДЕКС+ПОИСКПОЗ"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Найдите ответственного менеджера и тип канала по коду канала из листа «Данные»."
    ws["A2"].font = NOTE_FONT

    headers = ["Код канала", "Канал", "Тип канала", "Ответственный менеджер"]
    header_row = 4
    for j, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=j, value=h)
    style_header_row(ws, header_row, len(headers))

    rows = [
        ("YD", "Яндекс.Директ", "Платный", "Анна Смирнова"),
        ("VK", "VK Реклама", "Платный", "Дмитрий Орлов"),
        ("SEO", "SEO", "Условно бесплатный", "Мария Волкова"),
        ("EM", "Email", "Собственный", "Игорь Петров"),
        ("PT", "Партнёрка", "Заслуженный", "Ольга Смирнова"),
    ]
    for i, r in enumerate(rows, start=header_row + 1):
        for j, v in enumerate(r, start=1):
            ws.cell(row=i, column=j, value=v)

    autosize(ws, [12, 18, 20, 26])
    return ws


def build_exercises_sheet(wb):
    ws = wb.create_sheet("Задания")
    ws["A1"] = "Практические задания"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Выполняйте на листе «Данные» или в копии этого файла — место для формул есть в столбце C напротив каждого задания."
    ws["A2"].font = NOTE_FONT

    tasks = [
        ("Модуль 1. Формулы и ссылки", None),
        ("Посчитайте общий расход по всем кампаниям (СУММ)", ""),
        ("Посчитайте средний расход на одну кампанию (СРЗНАЧ)", ""),
        ("Посчитайте, сколько всего кампаний в таблице (СЧЁТ)", ""),
        ("Округлите средний CTR (клики/показы) до двух знаков после запятой", ""),
        ("Модуль 2. Текст, даты и поиск данных", None),
        ("С помощью ВПР найдите ответственного менеджера для кампании YD_202603_01", ""),
        ("С помощью ИНДЕКС+ПОИСКПОЗ найдите тип канала для VK Реклама", ""),
        ("Посчитайте, сколько дней прошло с самой ранней даты кампании до сегодняшнего дня", ""),
        ("Модуль 3. Логика, сортировка и фильтры", None),
        ("Посчитайте суммарный расход только по активным кампаниям (СУММЕСЛИ)", ""),
        ("Посчитайте количество кампаний по каналу «SEO» (СЧЁТЕСЛИ)", ""),
        ("С помощью ЕСЛИ пометьте кампании с расходом больше 10 000 ₽ как «Крупная»", ""),
        ("Отфильтруйте таблицу по каналу VK Реклама и статусу «Активна»", ""),
        ("Модуль 4. Сводные таблицы", None),
        ("Постройте сводную таблицу: сумма расхода по каждому каналу", ""),
        ("Добавьте в сводную таблицу количество заявок по каждому каналу и месяцу", ""),
        ("Модуль 5. Диаграммы", None),
        ("Постройте столбчатую диаграмму расходов по каналам", ""),
        ("Постройте линейный график заявок по месяцам", ""),
    ]

    row = 4
    for task, _ in tasks:
        cell = ws.cell(row=row, column=1, value=task)
        if _ is None:
            cell.font = Font(bold=True, color="E2532F")
        row += 1

    autosize(ws, [70, 6, 30])
    return ws


def main():
    wb = Workbook()
    build_data_sheet(wb)
    build_lookup_sheet(wb)
    build_exercises_sheet(wb)
    out_path = "/private/tmp/claude-501/-Users-igorshenshin-Developer-Web-cleanor-web/6514beb6-4747-44dc-be99-ab4c6125bf4a/scratchpad/183094-workbook/Excel-trenazher-183094.xlsx"
    wb.save(out_path)
    print("saved", out_path)


if __name__ == "__main__":
    main()

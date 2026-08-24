#!/usr/bin/env python3
"""Build the downloadable practice workbook for course 183077 (paid, tier-1
analytics course). Unlike the free Excel course's synthetic dataset, this one
is REAL cleanor.app GA4 data (property 528540338, pulled 2026-08-24) — the
actual continuation of the "/tools traffic dip" story from the free course
297305, now with August data through the 23rd instead of a partial snapshot.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BRAND_HEADER_FILL = PatternFill("solid", fgColor="E2532F")
BRAND_HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1C1A18")
NOTE_FONT = Font(italic=True, color="57524B")

# Real data, cleanor.app GA4, pulled 2026-08-24. August is partial (1-23).
SITE_TOTALS = [
    # month, channel, sessions, conversions, engagement_rate
    ("2026-05", "Organic Search", 146, 0, 0.034),
    ("2026-05", "Direct", 44, 0, 0.068),
    ("2026-05", "Referral", 11, 0, 0.0),
    ("2026-05", "Organic Social", 3, 0, 0.0),
    ("2026-05", "Unassigned", 2, 0, 0.0),
    ("2026-06", "Organic Search", 1286, 0, 0.598),
    ("2026-06", "Direct", 961, 0, 0.200),
    ("2026-06", "AI Assistant", 37, 0, 0.541),
    ("2026-06", "Organic Social", 20, 0, 0.600),
    ("2026-06", "Unassigned", 9, 0, 0.0),
    ("2026-06", "Referral", 6, 0, 0.667),
    ("2026-07", "Direct", 9512, 843, 0.204),
    ("2026-07", "Organic Search", 8666, 7440, 0.708),
    ("2026-07", "AI Assistant", 275, 196, 0.669),
    ("2026-07", "Unassigned", 192, 48, 0.448),
    ("2026-07", "Referral", 143, 61, 0.678),
    ("2026-07", "Organic Social", 32, 0, 0.719),
    ("2026-07", "Organic Video", 9, 3, 0.889),
    ("2026-07", "Paid Search", 6, 0, 0.0),
    ("2026-08*", "Organic Search", 10331, 8988, 0.681),
    ("2026-08*", "Direct", 9649, 1608, 0.150),
    ("2026-08*", "Unassigned", 767, 237, 0.061),
    ("2026-08*", "AI Assistant", 490, 495, 0.694),
    ("2026-08*", "Referral", 136, 142, 0.647),
    ("2026-08*", "Organic Social", 13, 0, 0.385),
    ("2026-08*", "Paid Search", 7, 0, 0.0),
    ("2026-08*", "Organic Video", 2, 0, 1.0),
]

TOOLS_PAGE = [
    # month, channel, sessions (exact page path /tools)
    ("2026-06", "Organic Search", 62),
    ("2026-06", "Direct", 12),
    ("2026-06", "Organic Social", 4),
    ("2026-06", "Referral", 1),
    ("2026-07", "Organic Search", 221),
    ("2026-07", "Direct", 75),
    ("2026-07", "Referral", 30),
    ("2026-07", "Unassigned", 24),
    ("2026-07", "AI Assistant", 11),
    ("2026-07", "Organic Social", 10),
    ("2026-07", "Organic Video", 2),
    ("2026-08*", "Organic Search", 221),
    ("2026-08*", "Direct", 80),
    ("2026-08*", "Unassigned", 32),
    ("2026-08*", "AI Assistant", 19),
    ("2026-08*", "Referral", 6),
]

TASKS = [
    ("Модуль 1. От вопроса к метрикам", None),
    ("Сформулируйте вопрос бизнеса по данным листа «/tools по каналам»: стоит ли беспокоиться о Referral-трафике?", ""),
    ("Выпишите 3 метрики, которые нужны, чтобы ответить на этот вопрос, и почему именно они", ""),
    ("Модуль 2. Сбор и очистка данных", None),
    ("Приведите август к сопоставимому виду с июлем: посчитайте дневной темп (сессии / число дней) для каждого канала", ""),
    ("Определите, какой вывод «в лоб» (без нормировки на число дней) был бы ошибочным и почему", ""),
    ("Модуль 3. Дашборд", None),
    ("Постройте сводную таблицу: сессии по каналу и месяцу для листа «/tools по каналам»", ""),
    ("Добавьте столбец с дневным темпом трафика (для сопоставимости неполного августа с июлем)", ""),
    ("Постройте диаграмму, которая наглядно показывает, какие каналы выросли, а какие просели", ""),
    ("Модуль 4. Отчёт с рекомендацией", None),
    ("Напишите вывод на 3-4 предложения: что реально произошло с трафиком /tools в августе", ""),
    ("Сформулируйте рекомендацию с конкретной цифрой: что делать с Referral-каналом", ""),
    ("Модуль 5. Защита решения", None),
    ("Запишите 3 вопроса, которые может задать руководитель, увидев вашу рекомендацию", ""),
    ("Подготовьте ответ на каждый, опираясь на цифры из дашборда, а не на предположения", ""),
]


def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = BRAND_HEADER_FILL
        c.font = BRAND_HEADER_FONT
        c.alignment = Alignment(horizontal="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_site_sheet(wb):
    ws = wb.active
    ws.title = "Сайт по каналам"
    ws["A1"] = "cleanor.app — реальные данные GA4 (property 528540338)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Выгружено 2026-08-24. Август отмечен * — данные только по 23 число включительно, не полный месяц."
    ws["A2"].font = NOTE_FONT

    headers = ["Месяц", "Канал", "Сеансы", "Конверсии", "Вовлечённость"]
    header_row = 4
    for j, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=j, value=h)
    style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    for month, channel, sessions, conv, eng in SITE_TOTALS:
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=channel)
        ws.cell(row=row, column=3, value=sessions)
        ws.cell(row=row, column=4, value=conv)
        ws.cell(row=row, column=5, value=round(eng, 3))
        row += 1

    autosize(ws, [12, 18, 10, 12, 14])
    ws.freeze_panes = "A5"
    return ws


def build_tools_sheet(wb):
    ws = wb.create_sheet("Данные по tools каналам")
    ws["A1"] = "Страница /tools — реальные сеансы по каналам и месяцам"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Тот же кейс, что в бесплатном курсе «Маркетинговая аналитика», но с полными данными за август."
    ws["A2"].font = NOTE_FONT

    headers = ["Месяц", "Канал", "Сеансы"]
    header_row = 4
    for j, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=j, value=h)
    style_header_row(ws, header_row, len(headers))

    row = header_row + 1
    for month, channel, sessions in TOOLS_PAGE:
        ws.cell(row=row, column=1, value=month)
        ws.cell(row=row, column=2, value=channel)
        ws.cell(row=row, column=3, value=sessions)
        row += 1

    autosize(ws, [12, 18, 10])
    ws.freeze_panes = "A5"
    return ws


def build_exercises_sheet(wb):
    ws = wb.create_sheet("Задания")
    ws["A1"] = "Практические задания курса"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Выполняйте на копии этого файла, используя листы «Сайт по каналам» и «Данные по /tools каналам»."
    ws["A2"].font = NOTE_FONT

    row = 4
    for task, _ in TASKS:
        cell = ws.cell(row=row, column=1, value=task)
        if _ is None:
            cell.font = Font(bold=True, color="E2532F")
        row += 1

    autosize(ws, [90])
    return ws


def main():
    wb = Workbook()
    build_site_sheet(wb)
    build_tools_sheet(wb)
    build_exercises_sheet(wb)
    out_path = "/private/tmp/claude-501/-Users-igorshenshin-Developer-Web-cleanor-web/6514beb6-4747-44dc-be99-ab4c6125bf4a/scratchpad/183077-workbook/Analitika-Praktika-183077.xlsx"
    wb.save(out_path)
    print("saved", out_path)


if __name__ == "__main__":
    main()
